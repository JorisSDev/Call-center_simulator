from collections import deque
import random
import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import xlsxwriter

class Caller:
    name = "empty"
    number = ""
    isVIP = -1
    reasonForCall = ""
    notes = ""
 
    def isCallerVIP(self):
        
        if self.isVIP == True:
            print("Client is VIP")
        else:
            print("Client is a normal client")
 
def ShowMainMenu():
    text = """
    -------------------------------------
    [0] :: Stop work
    [1] :: Still running(skip)
    [2] :: Chance for call
    [3] :: Accept call/Decline call 
    -------------------------------------
    """
    print(text)

def CallMenu(caller):
    print("-------------------------------------")
    print(caller.name,caller.number,caller.isVIP,caller.reasonForCall,caller.notes)
    print("-------------------------------------")
    global d
    global n
    
    while True:
        find = False
        x = input("Press [A] to accept call, [D] to decline: ")
        if   x == "A":
            while True:
                x = input("Press [E] to end call: ")
                if x == "E":
                    
                    path = "./2 kursas/Call_center/output.xlsx"

                    wb = load_workbook(path)
                    ws = wb.worksheets[0]

                    x=0
                    col = ["A","B","C","D","E"]
                    row = n
                    
                    for column in col:
                        y = str(row)
                        cell = column + str(row)
                        match column:
                            case "A":
                                ws[cell] = caller.name
                            case "B":
                                ws[cell] = caller.number
                            case "C":
                                ws[cell] = caller.isVIP
                            case "D":
                                ws[cell] = caller.reasonForCall
                            case "E":
                                ws[cell] = caller.notes
                            
                    wb.save(path)   
                    print("Call ended, data filled")
                    n+=1
                    d.popleft()
                    return 0          
        elif x=="D":
            print("Call was declined")
            d.popleft()
            return 0 
    
    
def GetCallerInfo(i:int):
    
    # Define variable to load the dataframe
    dataframeLoader = openpyxl.load_workbook("./2 kursas/Call_center/data.xlsx")
    
    # Define variable to read sheet
    dataframe = dataframeLoader.active
    
    global caller
    obj = dataframe.cell(row=i,column=1)
    caller.name = obj.value
    
    obj = dataframe.cell(row=i,column=2)
    caller.number = obj.value
    
    obj = dataframe.cell(row=i,column=3)
    caller.isVIP = int(obj.value)
    
    obj = dataframe.cell(row=i,column=4)
    caller.reasonForCall = obj.value
    
    obj = dataframe.cell(row=i,column=5)
    caller.notes = obj.value
    
    print(caller.number)

    
    
def ShowWaitList(d:deque):
    i=1
    text = """
    -------------------------------------
                WAITING LINE
    -------------------------------------
    """
    print(text)
    for x in d:
        print(i,x.number,x.reasonForCall,x.isVIP,x.notes)
        i+=1
    
    
# Driver code
# Object instantiation'

path = "./2 kursas/Call_center/output.xlsx"
if os.path.exists(path):
    os.remove(path)
    print("removed")

wb = openpyxl.Workbook()
wb.save(path)

global d; d = deque()
global n; n:int; n=1

i=1
state="off"

stateText="""
[on]   -> Start work
[exit] -> exit program
"""

#Start callCenter // regulated by state
while True:
    print(stateText)
    state="on" ;state = input("Input: ")
    if state == "on":
        while True:
            ShowWaitList(d)
            ShowMainMenu()
            select = input("Make your selection: ")
            
            if select == "0":
                print("breaking")
                break
            elif select == "1":
                print("Still running")
            elif select == "2":
                #new caller
                chance = random.randint(1,10)
                if chance<=5:
                    i += 1
                    print("New caller!")
                     
                    global caller
                    caller = Caller()
                    
                    GetCallerInfo(i)
    
                    if caller.isVIP == 1:
                        d.appendleft(caller)
                    else:
                        d.append(caller)
                else:
                    print("No calls recieved :((")
                    
            elif select == "3":
                CallMenu(d[0])
            else:
                print("ERROR wrong input!")
    elif state == "exit":
        break
    else:
        print("ERROR wrong input!")

